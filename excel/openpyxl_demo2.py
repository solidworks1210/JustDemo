# -*- coding:utf-8 -*- 
# --------------------
# Author:       SDN
# Description:  
# --------------------

from openpyxl import load_workbook
from openpyxl import Workbook

# 从现有文件获取workbook
workbook_from_file = load_workbook(filename='test_excel.xlsx')
# 获取该workbook中的所有worksheet的名字list
all_worksheet_name = workbook_from_file.sheetnames
print all_worksheet_name  # [u'Sheet1']
# 获取该workbook中的所有worksheet对象的list
all_worksheet_obj = workbook_from_file.worksheets
print all_worksheet_obj  # [<Worksheet "Sheet1">]
# 根据worksheet的名字获取worksheet对象，方法get_sheet_by_name已弃用
worksheet_obj = workbook_from_file['Sheet1']
print worksheet_obj  # <Worksheet "Sheet1">
# 获取worksheet中的所有行对象，列 columns
print worksheet_obj.rows  # <generator object _cells_by_row at 0x0000000004401F30>
for _ in worksheet_obj.rows:
    print _  # (<Cell u'Sheet1'.A1>, <Cell u'Sheet1'.B1>, <Cell u'Sheet1'.C1>)
# 获取单元格对象，index从1开始
print worksheet_obj.cell(1, 1)  # <Cell u'Sheet1'.A1>
# 获取单元格对象的值
print worksheet_obj.cell(1, 1).value    # 日期
print worksheet_obj.cell(2, 1).value    # 2018-08-23 00:00:00
print type(worksheet_obj.cell(2, 1).value)    # <type 'datetime.datetime'>
print type(worksheet_obj.cell(2, 2).value)    # <type 'unicode'>
print type(worksheet_obj.cell(2, 3).value)    # <type 'long'>
# 设置单元格的值
# worksheet_obj.cell(4, 1, '2018/8/26')
print type(worksheet_obj.cell(4, 1).value)
# workbook_from_file.save('test_excel.xlsx')
# 增加worksheet
ws_obj = workbook_from_file.create_sheet('hello_sheet', 1)
# workbook_from_file.save('test_excel.xlsx')