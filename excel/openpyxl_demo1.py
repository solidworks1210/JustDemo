# -*- coding:utf-8 -*- 
# --------------------
# Author:       SDN
# Description:  excel相关操作demo
# --------------------

filename = 'test_execl.xlsx'

"""
openpyxl:
该模块支持最新版的Excel文件格式，对Excel文件具有响应的读写操作，对此有专门的Reader和Writer两个类，便于对Excel文件的操作。虽然如此，
但我一般还是用默认的workbook来进行操作。
"""

from openpyxl.reader.excel import load_workbook # 这个方法用来打开已存在的excel文件

# 读取表格内容
wb = load_workbook(filename)
# 显示工作表的索引范围
wb.get_named_ranges()
# 显示所有工作表的名字
wb.get_sheet_names()
# 取得第一张表
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
# 获取表名
sheet_title = ws.title
# 获取表的行数
ws.get_highest_row()
# 获取表的列数
ws.get_highest_column()
# 单元格的读取，此处和xlrd的读取方式很相近，都是通过行和列的索引来读取
# 读取B1单元格中的内容
cell_value1 = ws.cell(0, 1).value
# 当然也支持通过Excel坐标来读取数据，代码如下
# 读取B1单元格中的内容
cell_value2 = ws.cell("B1").value


"""
写文件，只有一种操作方式，就是通过坐标。例如要向单元格C1写数据，就要用类似ws.cell(“C1”).value = something这样的方式。
一般推荐的方式是用openpyxl中的Writer类来实现。
"""

from openpyxl.workbook import Workbook  # 这个类用来生成一个excel文件

# ExcelWriter,里面封装好了对Excel的写操作
from openpyxl.writer.excel import ExcelWriter

# 新建一个workbook
wb = Workbook()

# 新建一个excelWriter
ew = ExcelWriter(workbook=wb)

# 第一个sheet是ws
ws = wb.worksheets[0]

# 设置ws的名称
ws.title = "range names"

# 向某个单元格中写入数据
ws.cell("C1").value = u'哈哈'

# 最后保存文件
ew.save(filename=r'empty_book.xlsx')

"""
向某个单元格内写文件时要先知道它对应的行数和列数，这里注意行数是从1开始计数的，而列则是从字母A开始，因此第一行第一列是A1，
这实际上是采用坐标方式操作Excel。例如，想向表格的第三行第一列插入一个数值1.2，
用xlwt写就是table.write(2, 0, 1.2), 因为xlwt中行列索引都从0开始；
而如果用openpyxl写就是ws.cell(“A3”).value = 1.2。
一般对于一个较大的列数，需要通过get_column_letter函数得到相应的字符，然后再调用cell函数写入。
下面是我之前写的一个代码的一部分，可以用来演示将多位数组保存到Excel文件中。为了体现多维数组，这里用到了numpy，另外这里为了简化过程，没有用ExcelWriter。代码如下：
"""

from openpyxl import Workbook
from openpyxl.cell import get_column_letter

import numpy as np

# 生成一个对角阵
a = np.diag([1, 2, 3, 4, 5])

# 新建一个工作簿
wb = Workbook()
# 使用当前激活的工作表（默认就是Excel中的第一张表）
ws = wb.active
# 下面是对a的遍历，注意cell中行和列从1开始，a中索引从0开始。
for row in xrange(1, a.shape[0] + 1):
    for col in xrange(1, a.shape[1] + 1):
        col_letter = get_column_letter(col)
        ws.cell('%s%s' % (col_letter, row)).value = a[row - 1, col - 1]
wb.save('test.xlsx')
